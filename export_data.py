#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
export_data.py
把 data/ 資料夾的所有 JSON 歷史數據匯出成：
  1. taifex_summary.xlsx（本機下載用）
  2. 上傳至 Google Sheets（需設定 credentials）

每日一列，含所有關鍵指標摘要。

用法：
  python export_data.py           # 只匯出 Excel
  python export_data.py --sheets  # Excel + 寫入 Google Sheets
"""

import json
import os
import sys
import argparse
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA_DIR    = "data"
OUTPUT_FILE = "taifex_summary.xlsx"

# ============================================================
# 讀取 JSON
# ============================================================

def load_json(filepath):
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def build_summary_df():
    futures_list = load_json(os.path.join(DATA_DIR, "futures.json"))
    options_list = load_json(os.path.join(DATA_DIR, "options_pc.json"))
    pc_list      = load_json(os.path.join(DATA_DIR, "pc_ratio.json"))
    oi_list      = load_json(os.path.join(DATA_DIR, "oi_strike.json"))
    stock_list   = load_json(os.path.join(DATA_DIR, "stock_net.json"))

    def to_dict(records, key="date"):
        return {r[key]: r for r in records if key in r}

    fut_map   = to_dict(futures_list)
    opt_map   = to_dict(options_list)
    pc_map    = to_dict(pc_list)
    oi_map    = to_dict(oi_list)
    stock_map = to_dict(stock_list)

    all_dates = sorted(set(
        list(fut_map.keys()) + list(opt_map.keys()) +
        list(pc_map.keys())  + list(stock_map.keys())
    ))

    rows = []
    prev_fut = None

    for d in all_dates:
        fut   = fut_map.get(d, {})
        opt   = opt_map.get(d, {})
        pc    = pc_map.get(d, {})
        oi    = oi_map.get(d, {})
        stock = stock_map.get(d, {})

        txf      = fut.get("txf", {}) or {}
        subtotal = fut.get("subtotal", {}) or {}
        prev_txf = (prev_fut or {}).get("txf", {}) or {}

        net_oi     = txf.get("net_oi")
        prev_net   = prev_txf.get("net_oi")
        net_oi_chg = (net_oi - prev_net) if net_oi is not None and prev_net is not None else None

        # ── OI 週選 / 月選
        weekly  = oi.get("weekly",  {}) or {}
        monthly = oi.get("monthly", {}) or {}
        signals = oi.get("signals", [])
        signal_types = "、".join(s.get("type", "") for s in signals) if signals else ""

        row = {
            # 基本
            "日期":                      d,

            # 台股期貨
            "期貨_外資淨多單(口)":        txf.get("long_oi"),
            "期貨_外資淨空單(口)":        txf.get("short_oi"),
            "期貨_外資淨未平倉(口)":      net_oi,
            "期貨_外資淨未平倉(千元)":    txf.get("net_oi_amount"),
            "期貨_淨未平倉變化(口)":      net_oi_chg,

            # 期貨小計
            "小計_外資淨未平倉(口)":      subtotal.get("net_oi"),
            "小計_外資淨未平倉(千元)":    subtotal.get("net_oi_amount"),

            # 選擇權
            "選擇權_外資Call淨未平倉":    opt.get("net_call_oi"),
            "選擇權_外資Put淨未平倉":     opt.get("net_put_oi"),

            # P/C Ratio
            "PC比率(%)":                 pc.get("pc_oi"),
            "Put未平倉量":               pc.get("put_oi"),
            "Call未平倉量":              pc.get("call_oi"),

            # 週選堡壘
            "週選_最大Call壓力":          weekly.get("max_call_strike"),
            "週選_最大Call壓力OI":        weekly.get("max_call_oi"),
            "週選_最大Put支撐":           weekly.get("max_put_strike"),
            "週選_最大Put支撐OI":         weekly.get("max_put_oi"),
            "週選_壓力帶低":              (weekly.get("call_range") or {}).get("low"),
            "週選_壓力帶高":              (weekly.get("call_range") or {}).get("high"),
            "週選_支撐帶低":              (weekly.get("put_range") or {}).get("low"),
            "週選_支撐帶高":              (weekly.get("put_range") or {}).get("high"),
            "週選_Call集中度(%)":         weekly.get("call_concentration"),
            "週選_Put集中度(%)":          weekly.get("put_concentration"),
            "週選_Call加碼最多履約價":    weekly.get("max_call_chg_strike"),
            "週選_Call加碼口數":          weekly.get("max_call_chg"),
            "週選_Put加碼最多履約價":     weekly.get("max_put_chg_strike"),
            "週選_Put加碼口數":           weekly.get("max_put_chg"),

            # 月選重鎮
            "月選_最大Call壓力":          monthly.get("max_call_strike"),
            "月選_最大Call壓力OI":        monthly.get("max_call_oi"),
            "月選_最大Put支撐":           monthly.get("max_put_strike"),
            "月選_最大Put支撐OI":         monthly.get("max_put_oi"),
            "月選_壓力帶低":              (monthly.get("call_range") or {}).get("low"),
            "月選_壓力帶高":              (monthly.get("call_range") or {}).get("high"),
            "月選_支撐帶低":              (monthly.get("put_range") or {}).get("low"),
            "月選_支撐帶高":              (monthly.get("put_range") or {}).get("high"),
            "月選_Call集中度(%)":         monthly.get("call_concentration"),
            "月選_Put集中度(%)":          monthly.get("put_concentration"),
            "月選_Call加碼最多履約價":    monthly.get("max_call_chg_strike"),
            "月選_Call加碼口數":          monthly.get("max_call_chg"),
            "月選_Put加碼最多履約價":     monthly.get("max_put_chg_strike"),
            "月選_Put加碼口數":           monthly.get("max_put_chg"),

            # 異常警報
            "異常訊號":                   signal_types,

            # 現貨
            "現貨_外資買賣超(億)":        stock.get("foreign"),
            "現貨_投信買賣超(億)":        stock.get("trust"),
            "現貨_自營商買賣超(億)":      stock.get("dealer_total"),
            "現貨_三大合計(億)":          stock.get("total"),
            "現貨_外資近5日累計(億)":     stock.get("foreign_5d"),
        }
        rows.append(row)
        prev_fut = fut

    return pd.DataFrame(rows)


# ============================================================
# 匯出 Excel
# ============================================================

def export_excel(df):
    try:
        # 先用 pandas 寫出基本資料
        df.to_excel(OUTPUT_FILE, index=False, sheet_name="每日摘要")
    except PermissionError:
        print(f"\n[錯誤] 無法寫入 {OUTPUT_FILE}！")
        print(f"請先關閉正在開啟中的 Excel 檔案，然後再重新執行程式。")
        sys.exit(1)

    # 再用 openpyxl 美化
    wb = load_workbook(OUTPUT_FILE)
    ws = wb["每日摘要"]

    # ── 顏色定義
    COLOR_HEADER    = "1E2333"   # 深藍灰（header 背景）
    COLOR_FUTURES   = "1A3A5C"   # 深藍（期貨區）
    COLOR_OPTIONS   = "1A3A3A"   # 深綠（選擇權區）
    COLOR_PC        = "2D1F4E"   # 深紫（P/C）
    COLOR_OI        = "2A2A1A"   # 深黃（OI）
    COLOR_STOCK     = "1A3A1A"   # 深綠（現貨）
    COLOR_POS       = "D4EDDA"   # 淡綠（正值）
    COLOR_NEG       = "F8D7DA"   # 淡紅（負值）

    # ── Header 格式
    header_font    = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border    = Border(
        left=Side(style="thin", color="2D3748"),
        right=Side(style="thin", color="2D3748"),
        top=Side(style="thin", color="2D3748"),
        bottom=Side(style="thin", color="2D3748"),
    )

    # 欄位分組對應顏色
    col_colors = {}
    for i, col in enumerate(df.columns, 1):
        if "日期" in col:
            col_colors[i] = COLOR_HEADER
        elif "期貨" in col:
            col_colors[i] = COLOR_FUTURES
        elif "小計" in col:
            col_colors[i] = COLOR_FUTURES
        elif "選擇權" in col:
            col_colors[i] = COLOR_OPTIONS
        elif "PC" in col or "Put未" in col or "Call未" in col:
            col_colors[i] = COLOR_PC
        elif "履約價" in col:
            col_colors[i] = COLOR_OI
        elif "現貨" in col:
            col_colors[i] = COLOR_STOCK
        else:
            col_colors[i] = COLOR_HEADER

    # ── 設定 header 列（第1列）
    ws.row_dimensions[1].height = 36
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font      = header_font
        cell.alignment = header_align
        cell.border    = thin_border
        cell.fill      = PatternFill("solid", start_color=col_colors.get(col_idx, COLOR_HEADER))

    # ── 資料列格式
    data_font  = Font(name="Arial", size=9)
    data_align = Alignment(horizontal="right", vertical="center")
    date_align = Alignment(horizontal="center", vertical="center")

    # 需要條件色的欄位（正值=淡綠、負值=淡紅）
    color_cols = {
        col: i+1 for i, col in enumerate(df.columns)
        if any(kw in col for kw in [
            "淨未平倉", "淨多單", "淨空單", "變化",
            "買賣超", "累計", "Call淨", "Put淨",
            "加碼口數", "集中度",   
        ]) 
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
        ws.row_dimensions[row_idx].height = 16
        for col_idx, cell in enumerate(row, 1):
            cell.font   = data_font
            cell.border = thin_border

            # 日期欄
            if col_idx == 1:
                cell.alignment = date_align
                cell.font      = Font(name="Arial", size=9, bold=True)
                continue

            cell.alignment = data_align

            # 數字格式
            if cell.value is None or cell.value == "":
                cell.value = None
                continue

            try:
                val = float(cell.value)
                # 條件色
                col_name = df.columns[col_idx - 1]
                if col_name in color_cols:
                    if val > 0:
                        cell.fill = PatternFill("solid", start_color=COLOR_POS)
                    elif val < 0:
                        cell.fill = PatternFill("solid", start_color=COLOR_NEG)
            except (TypeError, ValueError):
                pass

    # ── 欄寬
    col_widths = {1: 12}   # 日期
    for i in range(2, len(df.columns) + 1):
        col_widths[i] = 16

    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 凍結首列首欄
    ws.freeze_panes = "B2"

    # ── 加入說明工作表
    ws_note = wb.create_sheet("說明")
    notes = [
        ["欄位", "說明", "單位"],
        ["期貨_外資淨未平倉(口)", "外資台股期貨多方-空方未平倉口數，正=淨多，負=淨空", "口"],
        ["期貨_外資淨未平倉(千元)", "對應的契約金額", "千元"],
        ["期貨_淨未平倉變化(口)", "與前一交易日的口數差異，正=加碼多方，負=減碼或加空", "口"],
        ["小計_外資淨未平倉(口)", "所有期貨商品合計的外資淨未平倉", "口"],
        ["選擇權_外資Call淨未平倉", "外資買方Call-賣方Call，正=淨買Call看多", "口"],
        ["選擇權_外資Put淨未平倉", "外資買方Put-賣方Put，負=淨賣Put看多", "口"],
        ["PC比率(%)", "全市場Put未平倉/Call未平倉，<80%偏多，>120%偏空", "%"],
        ["最大Call壓力履約價", "OI最大的Call履約價，視為近期壓力區", "點"],
        ["最大Put支撐履約價", "OI最大的Put履約價，視為近期支撐區", "點"],
        ["Call/Put加碼最多履約價", "當日OI增加最多的履約價，代表主力佈局焦點", "點"],
        ["現貨_外資買賣超(億)", "外資及陸資（不含外資自營商）現貨買賣差額", "億元"],
        ["現貨_外資近5日累計(億)", "最近5個交易日外資現貨買賣超加總", "億元"],
        ["週選_最大Call壓力", "下次週選結算前最大Call未平倉履約價（壓力區）", "點"],
        ["週選_最大Put支撐",  "下次週選結算前最大Put未平倉履約價（支撐區）", "點"],
        ["週選_壓力帶低/高",  "Call OI 前5大履約價的範圍，整個壓力帶而非單點", "點"],
        ["週選_Call集中度",   "最大Call OI / 總Call OI，越高磁吸效應越強", "%"],
        ["月選_最大Call壓力", "本月月選最大Call未平倉履約價（中期壓力）", "點"],
        ["月選_最大Put支撐",  "本月月選最大Put未平倉履約價（中期支撐）", "點"],
        ["異常訊號",          "週選 vs 月選交叉分析結果（短空長多/短多長空/結算磁吸等）", "—"],
    ]
    for note_row in notes:
        ws_note.append(note_row)

    note_header_font = Font(name="Arial", bold=True, size=10)
    for cell in ws_note[1]:
        cell.font = note_header_font
        cell.fill = PatternFill("solid", start_color=COLOR_HEADER)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)

    ws_note.column_dimensions["A"].width = 28
    ws_note.column_dimensions["B"].width = 55
    ws_note.column_dimensions["C"].width = 10

    wb.save(OUTPUT_FILE)
    print(f"✓ Excel 匯出完成：{OUTPUT_FILE}（{len(df)} 筆資料）")


# ============================================================
# 寫入 Google Sheets
# ============================================================

def export_google_sheets(df):
    """
    需要先安裝：pip install gspread google-auth
    需要先設定 Google Service Account credentials
    參考：https://docs.gspread.org/en/latest/oauth2.html
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        print("  [錯誤] 請先安裝：pip install gspread google-auth")
        return

    # credentials.json 路徑（Google Service Account 金鑰）
    CREDS_FILE    = "credentials.json"
    # 你的 Google Sheet 名稱或 ID
    SHEET_NAME    = "外資期權籌碼追蹤"
    WORKSHEET     = "每日摘要"

    if not os.path.exists(CREDS_FILE):
        print(f"  [錯誤] 找不到 {CREDS_FILE}，請先設定 Google Service Account")
        print("  參考說明：https://docs.gspread.org/en/latest/oauth2.html")
        return

    scopes = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]

    creds  = Credentials.from_service_account_file(CREDS_FILE, scopes=scopes)
    client = gspread.authorize(creds)

    # 開啟或建立試算表
    try:
        sh = client.open(SHEET_NAME)
    except gspread.SpreadsheetNotFound:
        sh = client.create(SHEET_NAME)
        print(f"  建立新試算表：{SHEET_NAME}")

    # 開啟或建立工作表
    try:
        ws = sh.worksheet(WORKSHEET)
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = sh.add_worksheet(title=WORKSHEET, rows=2000, cols=len(df.columns))

    # 寫入 header
    headers = list(df.columns)
    ws.update([headers], "A1")

    # 寫入資料（None 換成空字串，避免 gspread 錯誤）
    data = df.where(pd.notna(df), "").values.tolist()
    if data:
        ws.update(data, f"A2")

    # 凍結首列
    ws.freeze(rows=1)

    # 格式化 header（底色深藍，白字）
    sh.batch_update({
        "requests": [
            {
                "repeatCell": {
                    "range": {
                        "sheetId": ws.id,
                        "startRowIndex": 0,
                        "endRowIndex": 1,
                    },
                    "cell": {
                        "userEnteredFormat": {
                            "backgroundColor": {"red": 0.12, "green": 0.14, "blue": 0.20},
                            "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True},
                            "horizontalAlignment": "CENTER",
                        }
                    },
                    "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)"
                }
            }
        ]
    })

    sheet_url = f"https://docs.google.com/spreadsheets/d/{sh.id}"
    print(f"✓ Google Sheets 更新完成：{sheet_url}")
    print(f"  工作表：{WORKSHEET}，共 {len(df)} 筆資料")


# ============================================================
# 主程式
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--sheets", action="store_true", help="同時寫入 Google Sheets")
    args = parser.parse_args()

    print("=== 歷史數據匯出 ===\n")

    df = build_summary_df()

    if df.empty:
        print("[警告] 沒有資料，請先執行 fetch_taifex.py 和 fetch_stock_net.py")
        sys.exit(1)

    print(f"資料範圍：{df['日期'].iloc[0]} ～ {df['日期'].iloc[-1]}，共 {len(df)} 筆\n")

    # 匯出 Excel
    export_excel(df)

    # 匯出 Google Sheets（需要 --sheets 參數）
    if args.sheets:
        print("\n寫入 Google Sheets...")
        export_google_sheets(df)
    else:
        print("\n提示：加上 --sheets 參數可同時寫入 Google Sheets")
        print("      python export_data.py --sheets")

    print("\n=== 完成 ===")